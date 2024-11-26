import openpyxl

# Creazione di un nuovo workbook e selezione del foglio attivo
wb = openpyxl.Workbook()
ws = wb.active

# Impostazione dei titoli delle colonne
ws['A1'] = 'Nome Componente'
ws['B1'] = 'Ingresso PLC'
ws['C1'] = 'Commento'

# Dati da inserire
dati = [
    ('Componente1', 'Ingresso1', 'Commento1'),
    ('Componente2', 'Ingresso2', 'Commento2'),
    ('Componente3', 'Ingresso3', 'Commento3')
]

# Inserimento dei dati nel foglio
for riga in dati:
    ws.append(riga)

# Salvataggio del file Excel
wb.save('dati_componenti.xlsx')